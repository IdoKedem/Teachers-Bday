from openpyxl import load_workbook
from collections import defaultdict


def fix_date(date: str) -> str:
    """
    changes (or not) and returns a date string to a string in format DD/MM
    :param date: a date string containing day and month separated by a '/'
    :return: date represented in DD/MM format
    """
    lst = date.split("/")  # a list with two items DD, MM
    lst = [("0" + i if len(i) == 1 else i) for i in lst]  # adds "0" to day or month if needed
    return "/".join(lst)

def load_database_dict():
    """
    Reads xlsx file and loads python dict
    :return: dict in format {date string : [teacher names]}
    """
    dates_names_dict = defaultdict(lambda: [])
    wb = load_workbook('Teachers-Bday.xlsx')
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        date = fix_date(ws["B" + str(row)].value)
        name = ws["A" + str(row)].value

        dates_names_dict[date].append(name)
    return dates_names_dict

def add_teacher_to_db(date: str, full_name: str) -> None:
    """
    writes a teacher's birthdate to the .xlsx file and to dict
    :param date: a date string containing day and month separated by a '/'
    :param full_name: full name of the teacher which will be added
    :return: None
    """
    wb = load_workbook('Teachers-Bday.xlsx')
    ws = wb.active
    next_row = ws.max_row + 1
    ws['A' + str(next_row)] = full_name
    ws['B' + str(next_row)] = fix_date(date)
    wb.save('Teachers-Bday.xlsx')

def search_teachers(search_input: str):
    """
    searches a teacher according to a given string
    :param search_input: name or part of teacher
    :return: a dict featuring all teachers after search
    """
    teachers_dict = load_database_dict()
    search_result_dict = defaultdict(lambda: [])

    for date, teacher_list in teachers_dict.items():
        for teacher in teacher_list:
            if search_input in teacher:
                search_result_dict[date].append(teacher)
    return search_result_dict

def is_duplicate_teacher(name: str) -> bool:
    """
    checks if a teacher already exists in database
    :param name: teacher's name
    :return: True if already exists, False otherwise
    """
    teachers_dict = load_database_dict()
    for date, teacher_list in teachers_dict.items():
        for teacher in teacher_list:
            if name == teacher:
                return True
    return False